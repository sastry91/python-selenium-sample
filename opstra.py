from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support.ui import Select
import openpyxl as xl
from openpyxl import Workbook
import time
from datetime import date
import datetime
from sys import argv


try:
    today = str(date.today())
    timeStamp = str(datetime.datetime.now())
    timeStamp = timeStamp.replace(":", "")
    print(today)  # '2017-12-26'
    print(timeStamp)
    excel_location = argv[1]
    sheet_name = argv[2]
    chrome_driver_path = argv[3]
    print(excel_location)
    wb = xl.load_workbook(excel_location)
    current_sheet = wb[sheet_name]
    my_stock_list = []
    for row in range(1, current_sheet.max_row + 1):
        current_cell = current_sheet.cell(row, 1)
        my_stock_list.append(current_cell.value)
    print(my_stock_list)
    wb.create_sheet(today)
    wb.save(excel_location)
# Specifying incognito mode as you launch your browser[OPTIONAL]
    option = webdriver.ChromeOptions()
    option.add_argument("--incognito")

# Create new Instance of Chrome in incognito mode
    browser = webdriver.Chrome(executable_path=chrome_driver_path, chrome_options=option)

# Go to desired website
    browser.get("http://opstra.definedge.com/openinterest")
    browser.implicitly_wait(4)
    time.sleep(15)
    browser.find_element_by_xpath("//div[@class='v-input__control']").click()
    wb2 = xl.load_workbook(excel_location)
    current_sheet = wb2[today]
    current_cell_header1 = current_sheet.cell(1, 1)
    current_cell_header1.value = 'Stock Symbol'
    current_cell_header2 = current_sheet.cell(1, 2)
    current_cell_header2.value = 'Expiry Date'
    current_cell_header3 = current_sheet.cell(1, 3)
    current_cell_header3.value = 'Spot price'
    current_cell_header4 = current_sheet.cell(1, 4)
    current_cell_header4.value = 'Futures price'
    current_cell_header5 = current_sheet.cell(1, 5)
    current_cell_header5.value = 'Lot size'
    current_cell_header6 = current_sheet.cell(1, 6)
    current_cell_header6.value = 'PCR'
    current_cell_header7 = current_sheet.cell(1, 7)
    current_cell_header7.value = 'MaxPain Strike'
    current_cell_header8 = current_sheet.cell(1, 8)
    current_cell_header8.value = 'Modified MaxPain'
    rowCounter=2
    for eachStock in my_stock_list:
        print(eachStock)
        time.sleep(2)
        stock_value = browser.find_element_by_xpath("//input[@aria-label='Select Ticker']")
        stock_value.clear()
        browser.implicitly_wait(1)
        stock_value.send_keys(eachStock)
        browser.implicitly_wait(2)
        stock_value.send_keys(u'\ue007')
        #browser.implicitly_wait(10)
        time.sleep(15)
        print(browser.find_element_by_xpath("//input[@aria-label='Select Expiry']").get_attribute('value'))
        current_cell = current_sheet.cell(rowCounter, 1)
        current_cell.value = eachStock
        current_cell2 = current_sheet.cell(rowCounter, 2)
        current_cell2.value = str(browser.find_element_by_xpath("//input[@aria-label='Select Expiry']").get_attribute('value'))
        all_spans = browser.find_elements_by_xpath("//span[@class='v-chip__content']")
        columnCounter = 3
        for eachSpan in all_spans:
            prices = str(eachSpan.text)
            price = prices.split(':')
            print(price)
            print(price[1].strip())
            current_cell3 = current_sheet.cell(rowCounter, columnCounter)
            current_cell3.value = price[1].strip()
            # price("price with trim --")
            time.sleep(3)
            columnCounter += 1
        rowCounter += 1
    wb2.save(excel_location)
except Exception as e:
    print(e)
finally:
    browser.close()

