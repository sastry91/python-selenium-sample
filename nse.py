from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support.ui import Select
import openpyxl as xl
from openpyxl import Workbook
import time
from datetime import date
import datetime
from sys import argv


try:
    today = str(date.today())
    timeStamp = str(datetime.datetime.now())
    timeStamp = timeStamp.replace(":", "")
    print(today)  # '2017-12-26'
    print(timeStamp)
    excel_location = argv[1]
    #excel_location ="opstra-stock.xlsx"
    sheet_name = argv[2]
    #sheet_name = "stock-sym"
    chrome_driver_path = argv[3]
    #chrome_driver_path = "C://Users//magundu//Documents//Python//Screen-Scraping//chromedriver"
    print(excel_location)
    wb = xl.load_workbook(excel_location)
    current_sheet = wb[sheet_name]
    my_stock_list = []
    for row in range(1, current_sheet.max_row + 1):
        current_cell = current_sheet.cell(row, 1)
        my_stock_list.append(current_cell.value)
    print(my_stock_list)
    wb.create_sheet(today)
    wb.save(excel_location)
# Specifying incognito mode as you launch your browser[OPTIONAL]
    option = webdriver.ChromeOptions()
    option.add_argument("--incognito")

# Create new Instance of Chrome in incognito mode
    browser = webdriver.Chrome(executable_path=chrome_driver_path, chrome_options=option)

# Go to desired website
    browser.get("https://www.nseindia.com/")
    time.sleep(4)
    wb2 = xl.load_workbook(excel_location)
    current_sheet = wb2[today]
    current_cell_header1 = current_sheet.cell(1, 1)
    current_cell_header1.value = 'Stock Symbol'
    current_cell_header2 = current_sheet.cell(1, 2)
    current_cell_header2.value = 'Date'
    current_cell_header3 = current_sheet.cell(1, 3)
    current_cell_header3.value = 'VWAP Value'
    rowCounter=2

    for eachStock in my_stock_list:
        try:
            print(eachStock)
            wait = WebDriverWait(browser, 10)
            adults_element = wait.until(EC.presence_of_element_located((By.ID, "QuoteSearch")))
            select = Select(adults_element)
            select.select_by_visible_text("Equity Derivatives")
            # time.sleep(10)
            # wait.until(EC.presence_of_element_located((By.XPATH,"//input[@class='arrow']")))
            stock_value = browser.find_element_by_id('fokeyword')
            # stock_value = browser.find_element_by_xpath("//input[@class='arrow']")
            stock_value.clear()
            browser.implicitly_wait(1)
            stock_value.send_keys(eachStock)
            time.sleep(2)
            stock_value.send_keys(u'\ue007')
            time.sleep(3)
            select_element = Select(browser.find_element_by_id("expiryDates"))
            print(select_element.options[2].text)
            select_element.select_by_visible_text(select_element.options[2].text)
            browser.find_element_by_css_selector('div>img').click()
            time.sleep(2)
            vwap_value = browser.find_element_by_id('vwap').text
            print(vwap_value)
            current_cell_header1 = current_sheet.cell(rowCounter, 1)
            current_cell_header1.value = eachStock
            current_cell_header1 = current_sheet.cell(rowCounter, 2)
            current_cell_header1.value = select_element.options[2].text
            current_cell_header1 = current_sheet.cell(rowCounter, 3)
            current_cell_header1.value = vwap_value
            rowCounter += 1
        except Exception as ex:
            print(ex)
            continue
    wb2.save(excel_location)
except Exception as e:
    print(e)
finally:
    wb2.save(excel_location)
    browser.close()

